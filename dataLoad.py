import os
import numpy as np

class dataLoad:
    #Private atribute
    __list_file = []

    #Constructor
    def __init__(self, path):
        self.path = path
        if os.path.isdir(path):
            self.__analyze_folder()
        elif os.path.isfile(path):
            file_type = self.__detect_file_type(path)
            self.__list_file = [{'Path': path, 'Type': file_type}]

    def get_list_type(self):
        return self.__list_file

    #Private functions
    def __analyze_folder(self):
        for filename in os.listdir(self.path):
            file_path = os.path.join(self.path, filename)
            if os.path.isfile(file_path):
                file_type = self.__detect_file_type(file_path)
                file = {'Path': file_path, 'Type': file_type}
                self.__list_file.append(file)

    def __detect_file_type(self, file_path):
        _, extension = os.path.splitext(file_path)
        return extension[1:]

    #Abstract functions
    def csv_read(self, file_path, type = "numpy") -> np.ndarray:
        pass

    def xlsx_read(self, file_path, type = "numpy") -> np.ndarray:
        pass

    def json_read(self, file_path, type = "numpy") -> np.ndarray:
        pass

    #Public functions
    def singleDatPack(self,file_path,file_type):
        datPack = []
        if file_type == "csv":
            datPack = self.csv_read(file_path)
        elif file_type == "xlsx":
            datPack = self.xlsx_read(file_path)
        elif file_type == "json":
            datPack = self.json_read(file_path)
        else:
            print("f{file_type} was not supported.")
        return datPack
    
    def throwDatPack(self):
        datPack = []
        if self.__list_file != None:
            if len(self.__list_file) == 1:
                datPack = self.singleDatPack(self.__list_file[0]["Path"],self.__list_file[0]["Type"])
            else:
                for item in self.__list_file:
                    data = self.singleDatPack(item['Path'],item['Type'])
                    datPack.append(data)
        return datPack
# Example usage
    # folder_path = "D:/my data/Medical_data/1_8_23_Nghia"
    # a = dataLoad(folder_path)
    # b = a.throwDatPack()
    # print(b)
    # print(a.get_list_type())

#Default setting

class datLoDe(dataLoad):
    def __init__(self,path):
        super().__init__(path)
    
    def csv_read(self, file_path, type="numpy"):
        import csv
        da = []
        with open(file_path, 'r') as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                tem =[]
                for element in row:
                    try:
                        tem.append(float(element))
                    except:
                        pass
                da.append(tem)
        return da
    
    def xlsx_read(self, file_path, type="numpy"):
        import openpyxl
        value = []
        dataframe = openpyxl.load_workbook(file_path)
        dataframe1 = dataframe.active
        for row in range(0, dataframe1.max_row):
            tem =[]
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                tem.append(col[row].value)
            value.append(tem)
        return value    
        

    
#Example script
# a = datLoDe("D:/my data/Medical_data/04_08_23_Long")
# b = a.throwDatPack()
# print(b)
